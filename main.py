import argparse
import hashlib
import json
import os
import random
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import chromadb
from llama_index.core import Document

from chunking import (
    build_vector_index,
    fixed_size_chunking,
    load_chunk_cache,
    save_chunk_cache,
    sentence_boundary_chunking,
)
from evaluation import RAGEvaluator
from llm import RAGLLM, complete_with_retry
from preprocess import load_or_create_preprocessed_documents, log_step
from query_rewrite import QueryRewriteRRFPipeline
from reranker import Reranker
from retrieval import BM25Retriever, HybridRetriever, SemanticRetriever


DEFAULT_API_KEY_ENV = "ZHIPUAI_API_KEY"
DEFAULT_GENERATION_MODEL = "glm-4.5-air"
DEFAULT_EVALUATION_MODEL = "glm-4-flash"
DEFAULT_REWRITE_MODEL = "glm-4-flash"
DEFAULT_EMBED_MODEL = "BAAI/bge-large-en-v1.5"

QUESTION_SETS = {
    "set1": "question/question_set1.json",
    "set2": "question/question_set2.json",
}


@dataclass(frozen=True)
class RAGConfig:
    name: str
    label: str
    chunking: str
    retrieval: str
    query_rewrite: bool
    description: str


RAG_CONFIGS: Dict[str, RAGConfig] = {
    "baseline": RAGConfig(
        name="baseline",
        label="Baseline Semantic RAG",
        chunking="fixed",
        retrieval="semantic",
        query_rewrite=False,
        description="fixed chunking + semantic retrieval",
    ),
    "advanced_chunking": RAGConfig(
        name="advanced_chunking",
        label="Baseline + Sentence-Boundary Chunking",
        chunking="advanced",
        retrieval="semantic",
        query_rewrite=False,
        description="sentence-boundary chunking + semantic retrieval",
    ),
    "query_rewrite": RAGConfig(
        name="query_rewrite",
        label="Baseline + Query Rewrite",
        chunking="fixed",
        retrieval="semantic",
        query_rewrite=True,
        description="fixed chunking + multi-query rewrite + semantic retrieval",
    ),
    "bm25": RAGConfig(
        name="bm25",
        label="Baseline + BM25",
        chunking="fixed",
        retrieval="bm25",
        query_rewrite=False,
        description="fixed chunking + BM25 retrieval",
    ),
    "hybrid": RAGConfig(
        name="hybrid",
        label="Baseline + Hybrid Search",
        chunking="fixed",
        retrieval="hybrid",
        query_rewrite=False,
        description="fixed chunking + semantic/BM25 weighted RRF",
    ),
    "full": RAGConfig(
        name="full",
        label="Full Upgraded RAG",
        chunking="advanced",
        retrieval="hybrid",
        query_rewrite=True,
        description="sentence-boundary chunking + multi-query rewrite + hybrid search",
    ),
}


def parse_config_names(config_arg: str) -> List[str]:
    if config_arg.strip().lower() == "all":
        return list(RAG_CONFIGS.keys())

    names = [name.strip() for name in config_arg.split(",") if name.strip()]
    unknown = [name for name in names if name not in RAG_CONFIGS]

    if unknown:
        valid = ", ".join(RAG_CONFIGS.keys())
        raise ValueError(f"Unknown config(s): {', '.join(unknown)}. Valid configs: {valid}, all")

    if not names:
        raise ValueError("At least one config must be selected.")

    return names


def collection_name_for(
    chunking_strategy: str,
    chunk_size: int,
    chunk_overlap: int,
    advanced_min_chunk_size: int,
    advanced_max_chunk_size: int,
    advanced_sentence_overlap: int,
) -> str:
    if chunking_strategy == "fixed":
        return f"rag_fixed_{chunk_size}_{chunk_overlap}"

    return (
        f"rag_sentence_{advanced_min_chunk_size}_"
        f"{advanced_max_chunk_size}_o{advanced_sentence_overlap}"
    )


def chunk_cache_path_for(
    cache_dir: str,
    chunking_strategy: str,
    chunk_size: int,
    chunk_overlap: int,
    advanced_min_chunk_size: int,
    advanced_max_chunk_size: int,
    advanced_sentence_overlap: int,
) -> str:
    if chunking_strategy == "fixed":
        file_name = f"chunks_fixed_{chunk_size}_{chunk_overlap}.json"
    else:
        file_name = (
            f"chunks_sentence_min{advanced_min_chunk_size}_"
            f"max{advanced_max_chunk_size}_overlap{advanced_sentence_overlap}.json"
        )

    return str(Path(cache_dir) / file_name)


def load_questions(path: str, question_set_name: Optional[str] = None) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as file:
        questions = json.load(file)

    if not isinstance(questions, list):
        raise ValueError("questions file must contain a JSON list.")

    set_name = question_set_name or Path(path).stem

    for index, question in enumerate(questions, start=1):
        question["question_set"] = set_name
        question["_sample_key"] = f"{set_name}:{question.get('question_id', index)}"

    return questions


def resolve_question_paths(question_set: str, questions_path: Optional[str]) -> List[tuple[str, str]]:
    if questions_path:
        return [(Path(questions_path).stem, questions_path)]

    question_set = question_set.strip()

    if question_set.lower() == "all":
        return list(QUESTION_SETS.items())

    set_names = []
    for name in question_set.split(","):
        normalized = name.strip()
        if normalized and normalized not in set_names:
            set_names.append(normalized)

    unknown_sets = [name for name in set_names if name not in QUESTION_SETS]

    if unknown_sets:
        valid_sets = ", ".join(list(QUESTION_SETS) + ["all"])
        raise ValueError(f"Unknown question set(s): {', '.join(unknown_sets)}. Valid question sets: {valid_sets}")

    if not set_names:
        raise ValueError("At least one question set must be selected.")

    return [(name, QUESTION_SETS[name]) for name in set_names]


def load_questions_from_paths(question_paths: List[tuple[str, str]]) -> List[Dict[str, Any]]:
    all_questions = []

    for question_set_name, path in question_paths:
        loaded_questions = load_questions(path, question_set_name=question_set_name)
        all_questions.extend(loaded_questions)

    return all_questions


def allocate_stratified_sample_counts(
    strata_sizes: Dict[str, int],
    sample_size: int,
) -> Dict[str, int]:
    total = sum(strata_sizes.values())

    if total <= 0 or sample_size <= 0:
        return {name: 0 for name in strata_sizes}

    sample_size = min(sample_size, total)
    allocations: Dict[str, int] = {}
    remainders = []

    for name, size in strata_sizes.items():
        exact = sample_size * (size / total)
        count = int(exact)
        allocations[name] = min(count, size)
        remainders.append((exact - count, name))

    remaining = sample_size - sum(allocations.values())

    for _, name in sorted(remainders, key=lambda item: (-item[0], item[1])):
        if remaining <= 0:
            break

        if allocations[name] < strata_sizes[name]:
            allocations[name] += 1
            remaining -= 1

    while remaining > 0:
        changed = False

        for name in sorted(strata_sizes):
            if allocations[name] < strata_sizes[name]:
                allocations[name] += 1
                remaining -= 1
                changed = True

                if remaining <= 0:
                    break

        if not changed:
            break

    return allocations


def sample_questions(
    questions: List[Dict[str, Any]],
    sample_size: Optional[int],
    seed: int,
) -> List[Dict[str, Any]]:
    if sample_size is None or sample_size <= 0:
        return list(questions)

    sample_size = min(sample_size, len(questions))

    if sample_size >= len(questions):
        return list(questions)

    rng = random.Random(seed)
    grouped_questions: Dict[str, List[Dict[str, Any]]] = {}

    for question in questions:
        question_set = str(question.get("question_set") or "default")
        grouped_questions.setdefault(question_set, []).append(question)

    strata_sizes = {
        name: len(items)
        for name, items in grouped_questions.items()
    }
    allocations = allocate_stratified_sample_counts(
        strata_sizes=strata_sizes,
        sample_size=sample_size,
    )

    selected = []

    for question_set, items in grouped_questions.items():
        count = allocations.get(question_set, 0)

        if count <= 0:
            continue

        selected.extend(rng.sample(items, count))

    return sorted(selected, key=lambda item: item.get("_sample_key", ""))


def question_set_counts(questions: List[Dict[str, Any]]) -> Dict[str, int]:
    counts: Dict[str, int] = {}

    for question in questions:
        question_set = str(question.get("question_set") or "default")
        counts[question_set] = counts.get(question_set, 0) + 1

    return dict(sorted(counts.items()))


def build_run_signature(
    args: argparse.Namespace,
    config_names: List[str],
    selected_questions: List[Dict[str, Any]],
    question_paths: List[tuple[str, str]],
) -> str:
    signature_data = {
        "configs": config_names,
        "question_paths": question_paths,
        "sample_keys": [question.get("_sample_key") for question in selected_questions],
        "generation_model": args.generation_model,
        "evaluation_model": args.evaluation_model,
        "rewrite_model": args.rewrite_model,
        "embed_model": args.embed_model,
        "chunk_size": args.chunk_size,
        "chunk_overlap": args.chunk_overlap,
        "advanced_min_chunk_size": args.advanced_min_chunk_size,
        "advanced_max_chunk_size": args.advanced_max_chunk_size,
        "advanced_sentence_overlap": args.advanced_sentence_overlap,
        "top_k_retrieve": args.top_k_retrieve,
        "top_k_rerank": args.top_k_rerank,
        "evaluation_k": args.evaluation_k,
        "rewrite_count": args.rewrite_count,
        "rewrite_top_k_per_query": args.rewrite_top_k_per_query,
        "hybrid_rrf_k": args.hybrid_rrf_k,
        "semantic_weight": args.semantic_weight,
        "bm25_weight": args.bm25_weight,
        "max_context_chars": args.max_context_chars,
    }
    raw = json.dumps(signature_data, ensure_ascii=False, sort_keys=True)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def load_checkpoint_results(
    checkpoint_path: str,
    run_signature: str,
) -> Dict[str, List[Dict[str, Any]]]:
    path = Path(checkpoint_path)

    if not path.exists():
        return {}

    try:
        with open(path, "r", encoding="utf-8") as file:
            checkpoint = json.load(file)
    except Exception as exc:
        log_step(f"Could not load checkpoint {checkpoint_path}: {exc}. Starting fresh.")
        return {}

    if checkpoint.get("run_signature") != run_signature:
        log_step(f"Checkpoint signature mismatch: {checkpoint_path}. Starting fresh.")
        return {}

    results = checkpoint.get("results_by_config", {})

    if not isinstance(results, dict):
        return {}

    log_step(f"Loaded checkpoint: {checkpoint_path}")
    return {
        str(config_name): config_results
        for config_name, config_results in results.items()
        if isinstance(config_results, list)
    }


def save_checkpoint_results(
    checkpoint_path: str,
    run_signature: str,
    results_by_config: Dict[str, List[Dict[str, Any]]],
) -> None:
    path = Path(checkpoint_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")

    checkpoint = {
        "run_signature": run_signature,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "results_by_config": results_by_config,
    }

    with open(tmp_path, "w", encoding="utf-8") as file:
        json.dump(checkpoint, file, ensure_ascii=False, indent=2)

    os.replace(tmp_path, path)


def chunk_documents(
    documents: List[Document],
    chunking_strategy: str,
    chunk_size: int,
    chunk_overlap: int,
    advanced_min_chunk_size: int,
    advanced_max_chunk_size: int,
    advanced_sentence_overlap: int,
    embed_model_name: str,
) -> List[Document]:
    if chunking_strategy == "fixed":
        return fixed_size_chunking(
            documents=documents,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    if chunking_strategy == "advanced":
        return sentence_boundary_chunking(
            documents=documents,
            min_chunk_size=advanced_min_chunk_size,
            max_chunk_size=advanced_max_chunk_size,
            sentence_overlap=advanced_sentence_overlap,
            semantic_model_name=embed_model_name,
        )

    raise ValueError(f"Unknown chunking strategy: {chunking_strategy}")


def get_chroma_collection_count(collection_name: str, chroma_dir: str) -> Optional[int]:
    client = chromadb.PersistentClient(path=chroma_dir)

    try:
        collection = client.get_collection(collection_name)
    except Exception:
        return None

    return collection.count()


def ensure_vector_collection(
    collection_name: str,
    chunks: List[Document],
    chroma_dir: str,
    embed_model_name: str,
    rebuild: bool,
) -> None:
    chunk_count = len(chunks)

    if not rebuild:
        vector_count = get_chroma_collection_count(collection_name, chroma_dir)

        if vector_count is None:
            log_step(f"Chroma collection not found: {collection_name}. Building collection.")
        elif vector_count == chunk_count:
            log_step(
                f"Reusing Chroma collection: {collection_name}, "
                f"chunks={chunk_count}, vectors={vector_count}"
            )
            return
        else:
            log_step(
                f"Vector count mismatch: chunks={chunk_count}, vectors={vector_count}. "
                "Rebuilding collection."
            )
    else:
        log_step(f"--rebuild-index set. Rebuilding Chroma collection: {collection_name}")

    if chunk_count == 0:
        raise ValueError(f"No chunks available for collection: {collection_name}")

    log_step(f"Embedding chunks into Chroma: {collection_name}, chunks={chunk_count}")

    build_vector_index(
        documents=chunks,
        collection_name=collection_name,
        chunking_strategy=None,
        embed_model_name=embed_model_name,
        chroma_dir=chroma_dir,
        reset_collection=True,
    )

    vector_count = get_chroma_collection_count(collection_name, chroma_dir)
    log_step(
        f"Chroma collection ready: {collection_name}, "
        f"chunks={chunk_count}, vectors={vector_count}"
    )


def load_or_create_chunks(
    documents: List[Document],
    chunking_strategy: str,
    cache_path: str,
    force_rebuild: bool,
    chunk_size: int,
    chunk_overlap: int,
    advanced_min_chunk_size: int,
    advanced_max_chunk_size: int,
    advanced_sentence_overlap: int,
    embed_model_name: str,
) -> List[Document]:
    if not force_rebuild and os.path.exists(cache_path):
        return load_chunk_cache(cache_path)

    if force_rebuild:
        log_step(f"Regenerating chunk cache: {cache_path}")
    else:
        log_step(f"Chunk cache not found. Creating: {cache_path}")

    chunks = chunk_documents(
        documents=documents,
        chunking_strategy=chunking_strategy,
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        advanced_min_chunk_size=advanced_min_chunk_size,
        advanced_max_chunk_size=advanced_max_chunk_size,
        advanced_sentence_overlap=advanced_sentence_overlap,
        embed_model_name=embed_model_name,
    )
    save_chunk_cache(chunks, cache_path)
    return chunks


def prepare_chunks_and_indexes(
    selected_configs: List[RAGConfig],
    pdf_dir: str,
    cache_dir: str,
    preprocess_cache: str,
    reprocess_pdf: bool,
    chroma_dir: str,
    embed_model_name: str,
    chunk_size: int,
    chunk_overlap: int,
    advanced_min_chunk_size: int,
    advanced_max_chunk_size: int,
    advanced_sentence_overlap: int,
    rebuild: bool,
) -> Dict[str, Dict[str, Any]]:
    log_step("Loading preprocessed PDFs")
    os.makedirs(cache_dir, exist_ok=True)
    documents = load_or_create_preprocessed_documents(
        pdf_dir=pdf_dir,
        cache_path=preprocess_cache,
        reprocess_pdf=reprocess_pdf,
    )

    requested_chunking_strategies = {config.chunking for config in selected_configs}
    chunking_strategies = [
        strategy
        for strategy in ("fixed", "advanced")
        if strategy in requested_chunking_strategies
    ]
    prepared: Dict[str, Dict[str, Any]] = {}

    for strategy in chunking_strategies:
        chunk_cache_path = chunk_cache_path_for(
            cache_dir=cache_dir,
            chunking_strategy=strategy,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            advanced_min_chunk_size=advanced_min_chunk_size,
            advanced_max_chunk_size=advanced_max_chunk_size,
            advanced_sentence_overlap=advanced_sentence_overlap,
        )
        chunks = load_or_create_chunks(
            documents=documents,
            chunking_strategy=strategy,
            cache_path=chunk_cache_path,
            force_rebuild=(reprocess_pdf or rebuild),
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            advanced_min_chunk_size=advanced_min_chunk_size,
            advanced_max_chunk_size=advanced_max_chunk_size,
            advanced_sentence_overlap=advanced_sentence_overlap,
            embed_model_name=embed_model_name,
        )

        collection_name = collection_name_for(
            chunking_strategy=strategy,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            advanced_min_chunk_size=advanced_min_chunk_size,
            advanced_max_chunk_size=advanced_max_chunk_size,
            advanced_sentence_overlap=advanced_sentence_overlap,
        )
        prepared[strategy] = {
            "chunks": chunks,
            "collection_name": collection_name,
            "chunk_cache_path": chunk_cache_path,
        }

    strategies_requiring_vectors = {
        config.chunking
        for config in selected_configs
        if config.retrieval in {"semantic", "hybrid"}
    }

    for strategy in strategies_requiring_vectors:
        ensure_vector_collection(
            collection_name=prepared[strategy]["collection_name"],
            chunks=prepared[strategy]["chunks"],
            chroma_dir=chroma_dir,
            embed_model_name=embed_model_name,
            rebuild=(rebuild or reprocess_pdf),
        )

    return prepared


def build_retriever(
    config: RAGConfig,
    prepared_entry: Dict[str, Any],
    chroma_dir: str,
    embed_model_name: str,
    hybrid_rrf_k: int,
    semantic_weight: float,
    bm25_weight: float,
    component_cache: Dict[str, Any],
):
    chunking_key = config.chunking

    def get_bm25():
        cache_key = f"bm25::{chunking_key}"
        if cache_key not in component_cache:
            component_cache[cache_key] = BM25Retriever(documents=prepared_entry["chunks"])
        return component_cache[cache_key]

    def get_semantic():
        cache_key = f"semantic::{chunking_key}"
        if cache_key not in component_cache:
            component_cache[cache_key] = SemanticRetriever(
                collection_name=prepared_entry["collection_name"],
                chroma_dir=chroma_dir,
                embed_model_name=embed_model_name,
            )
        return component_cache[cache_key]

    if config.retrieval == "bm25":
        return get_bm25()

    if config.retrieval == "semantic":
        return get_semantic()

    if config.retrieval == "hybrid":
        cache_key = (
            f"hybrid::{chunking_key}::{hybrid_rrf_k}::"
            f"{semantic_weight}::{bm25_weight}"
        )
        if cache_key not in component_cache:
            component_cache[cache_key] = HybridRetriever(
                semantic_retriever=get_semantic(),
                bm25_retriever=get_bm25(),
                rrf_k=hybrid_rrf_k,
                semantic_weight=semantic_weight,
                bm25_weight=bm25_weight,
            )
        return component_cache[cache_key]

    raise ValueError(f"Unknown retrieval strategy: {config.retrieval}")


def retrieve_documents(
    query: str,
    config: RAGConfig,
    retriever,
    component_cache: Dict[str, Any],
    api_key: Optional[str],
    rewrite_model: str,
    top_k_retrieve: int,
    rewrite_count: int,
    rewrite_top_k_per_query: int,
) -> Dict[str, Any]:
    if not config.query_rewrite:
        return {
            "effective_query": query,
            "query_type": None,
            "retrieval_mode": config.retrieval,
            "rewritten_queries": [],
            "retrieved_docs": retriever.retrieve(query, top_k=top_k_retrieve),
        }

    cache_key = (
        f"query_rewrite::{config.name}::{id(retriever)}::"
        f"{rewrite_model}::{rewrite_count}::{rewrite_top_k_per_query}::{top_k_retrieve}"
    )

    if cache_key not in component_cache:
        component_cache[cache_key] = QueryRewriteRRFPipeline(
            retriever=retriever,
            api_key=api_key,
            rewrite_model=rewrite_model,
            num_rewrites=rewrite_count,
            top_k_per_query=rewrite_top_k_per_query,
            final_top_k=top_k_retrieve,
        )

    rewrite_pipeline = component_cache[cache_key]
    result = rewrite_pipeline.retrieve_with_queries(query)

    return {
        "effective_query": query,
        "query_type": result["query_type"],
        "retrieval_mode": "llm_only_chat" if result["query_type"] == "general_chat" else "query_rewrite_rrf",
        "rewritten_queries": result["rewritten_queries"],
        "retrieved_docs": result["candidate_docs"],
    }


def generate_chat_answer(generator: RAGLLM, query: str) -> str:
    prompt = f"""
You are a helpful assistant.

The user query is classified as general chat, so do not use document retrieval.
Answer naturally and directly.

User query:
{query}

Answer:
"""
    response = complete_with_retry(generator.llm, prompt)
    return response.text.strip()


def skipped_chat_evaluation(query: str, k: int) -> Dict[str, Any]:
    return {
        "query": query,
        f"Page Recall@{k}": 0.0,
        f"Answer-Supported Recall@{k}": 0.0,
        f"Precision@{k}": 0.0,
        f"Precision Scores@{k}": [],
        "Groundedness": 1.0,
        "evaluation_skipped": True,
        "skip_reason": "general_chat_llm_only",
    }


def doc_source_key(doc: Dict[str, Any]) -> str:
    metadata = doc.get("metadata", {}) or {}
    return str(
        metadata.get("file_name")
        or metadata.get("source_name")
        or metadata.get("source")
        or metadata.get("document")
        or "unknown_source"
    )


def diversify_comparison_sources(
    reranked_docs: List[Dict[str, Any]],
    top_k: int,
    min_sources: int = 2,
) -> List[Dict[str, Any]]:
    if top_k <= 0 or not reranked_docs:
        return []

    available_sources = []
    seen_sources = set()

    for doc in reranked_docs:
        source = doc_source_key(doc)
        if source not in seen_sources:
            available_sources.append(source)
            seen_sources.add(source)

    required_sources = min(min_sources, top_k, len(available_sources))

    if required_sources <= 1:
        return reranked_docs[:top_k]

    selected = []
    selected_ids = set()

    for source in available_sources[:required_sources]:
        for doc in reranked_docs:
            doc_id = str(doc.get("id") or id(doc))
            if doc_id in selected_ids:
                continue
            if doc_source_key(doc) == source:
                selected.append(doc)
                selected_ids.add(doc_id)
                break

    for doc in reranked_docs:
        if len(selected) >= top_k:
            break

        doc_id = str(doc.get("id") or id(doc))
        if doc_id in selected_ids:
            continue

        selected.append(doc)
        selected_ids.add(doc_id)

    return selected[:top_k]


def average_metric(results: List[Dict[str, Any]], metric_name: str) -> float:
    if not results:
        return 0.0

    return sum(float(result["evaluation"].get(metric_name, 0.0)) for result in results) / len(results)


def summarize_config_results(config_name: str, results: List[Dict[str, Any]], k: int) -> Dict[str, Any]:
    return {
        "config": config_name,
        "questions": len(results),
        f"Page Recall@{k}": average_metric(results, f"Page Recall@{k}"),
        f"Answer-Supported Recall@{k}": average_metric(results, f"Answer-Supported Recall@{k}"),
        f"Precision@{k}": average_metric(results, f"Precision@{k}"),
        "Groundedness": average_metric(results, "Groundedness"),
    }


def run_one_config(
    config: RAGConfig,
    questions: List[Dict[str, Any]],
    prepared_entry: Dict[str, Any],
    component_cache: Dict[str, Any],
    reranker: Reranker,
    api_key: Optional[str],
    generation_model: str,
    evaluation_model: str,
    rewrite_model: str,
    embed_model_name: str,
    chroma_dir: str,
    top_k_retrieve: int,
    top_k_rerank: int,
    evaluation_k: int,
    rewrite_count: int,
    rewrite_top_k_per_query: int,
    hybrid_rrf_k: int,
    semantic_weight: float,
    bm25_weight: float,
    max_context_chars: Optional[int],
    existing_results: Optional[List[Dict[str, Any]]] = None,
    checkpoint_callback: Optional[Callable[[str, List[Dict[str, Any]]], None]] = None,
) -> Dict[str, Any]:
    log_step(f"Running config: {config.name} ({config.description})")

    retriever = build_retriever(
        config=config,
        prepared_entry=prepared_entry,
        chroma_dir=chroma_dir,
        embed_model_name=embed_model_name,
        hybrid_rrf_k=hybrid_rrf_k,
        semantic_weight=semantic_weight,
        bm25_weight=bm25_weight,
        component_cache=component_cache,
    )
    generator = RAGLLM(api_key=api_key, model=generation_model)
    evaluator = RAGEvaluator(api_key=api_key, model=evaluation_model)

    question_order = [
        question.get("_sample_key") or str(question.get("question_id"))
        for question in questions
    ]
    question_results_by_key: Dict[str, Dict[str, Any]] = {}

    for result in existing_results or []:
        result_key = result.get("sample_key") or str(result.get("question_id"))
        if result_key in question_order and result_key not in question_results_by_key:
            question_results_by_key[result_key] = result

    if question_results_by_key:
        log_step(
            f"{config.name}: resuming {len(question_results_by_key)}/"
            f"{len(questions)} completed questions from checkpoint"
        )

    for index, question in enumerate(questions, start=1):
        question_key = question.get("_sample_key") or str(question.get("question_id"))

        if question_key in question_results_by_key:
            log_step(f"{config.name}: question {index}/{len(questions)} already done, skipping")
            continue

        query = question["question"]
        log_step(f"{config.name}: question {index}/{len(questions)}")

        retrieval_result = retrieve_documents(
            query=query,
            config=config,
            retriever=retriever,
            component_cache=component_cache,
            api_key=api_key,
            rewrite_model=rewrite_model,
            top_k_retrieve=top_k_retrieve,
            rewrite_count=rewrite_count,
            rewrite_top_k_per_query=rewrite_top_k_per_query,
        )

        query_type = retrieval_result.get("query_type")

        if query_type == "general_chat":
            reranked_docs = []
            context = ""
            answer = generate_chat_answer(generator=generator, query=query)
        else:
            rerank_pool_size = (
                len(retrieval_result["retrieved_docs"])
                if query_type == "comparison"
                else top_k_rerank
            )
            reranked_pool = reranker.rerank(
                query=query,
                documents=retrieval_result["retrieved_docs"],
                top_k=rerank_pool_size,
            )
            if query_type == "comparison":
                reranked_docs = diversify_comparison_sources(
                    reranked_docs=reranked_pool,
                    top_k=top_k_rerank,
                )
            else:
                reranked_docs = reranked_pool[:top_k_rerank]
            context = generator.build_context(reranked_docs, max_context_chars=max_context_chars)
            answer = generator.generate_answer_from_context(query=query, context=context)

        if config.query_rewrite:
            answer = f"[Query Type: {query_type or 'general_rag'}]\n\n{answer}"

        result_for_eval = {
            "retrieved_docs": retrieval_result["retrieved_docs"],
            "reranked_docs": reranked_docs,
            "context": context,
            "answer": answer,
        }
        if query_type == "general_chat":
            evaluation = skipped_chat_evaluation(query=query, k=evaluation_k)
        else:
            evaluation = evaluator.evaluate_single(
                query=query,
                result=result_for_eval,
                ground_truth=question,
                k=evaluation_k,
            )

        result_record = {
            "question_id": question.get("question_id"),
            "question_set": question.get("question_set"),
            "sample_key": question.get("_sample_key"),
            "question": query,
            "standard_answer": question.get("answer", ""),
            "evidence": question.get("evidence", []),
            "query_type": query_type,
            "retrieval_mode": retrieval_result.get("retrieval_mode"),
            "rewritten_queries": retrieval_result["rewritten_queries"],
            "retrieved_docs": retrieval_result["retrieved_docs"],
            "reranked_docs": reranked_docs,
            "context": context,
            "answer": answer,
            "evaluation": evaluation,
        }
        question_results_by_key[question_key] = result_record

        if checkpoint_callback:
            ordered_checkpoint_results = [
                question_results_by_key[key]
                for key in question_order
                if key in question_results_by_key
            ]
            checkpoint_callback(config.name, ordered_checkpoint_results)

    question_results = [
        question_results_by_key[key]
        for key in question_order
        if key in question_results_by_key
    ]

    summary = summarize_config_results(config.name, question_results, evaluation_k)

    return {
        "config": config,
        "summary": summary,
        "results": question_results,
    }


def source_label(metadata: Dict[str, Any]) -> str:
    source = metadata.get("file_name") or metadata.get("source") or "Unknown source"
    page_start = metadata.get("page_start") or metadata.get("page_number")
    page_end = metadata.get("page_end") or page_start
    heading = metadata.get("nearby_heading")

    if page_start is None:
        label = str(source)
    elif page_start == page_end:
        label = f"{source}, page {page_start}"
    else:
        label = f"{source}, pages {page_start}-{page_end}"

    if heading:
        label += f", section: {heading}"

    return label


def doc_score(doc: Dict[str, Any]) -> str:
    parts = []

    if doc.get("retrieval_score") is not None:
        parts.append(f"retrieval={float(doc['retrieval_score']):.4f}")

    if doc.get("rerank_score") is not None:
        parts.append(f"rerank={float(doc['rerank_score']):.4f}")

    if doc.get("source_scores"):
        source_scores = ", ".join(
            f"{key}:{float(value):.4f}"
            for key, value in doc["source_scores"].items()
            if isinstance(value, (int, float))
        )
        if source_scores:
            parts.append(f"sources=({source_scores})")

    return "; ".join(parts) if parts else "score unavailable"


def truncate_text(text: str, max_chars: int) -> str:
    text = " ".join(str(text).split())
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3] + "..."


def format_evidence(evidence: List[Dict[str, Any]]) -> str:
    if not evidence:
        return "No citation provided."

    return "; ".join(
        f"{item.get('source', 'Unknown source')}, pages {item.get('pages', 'Unknown')}"
        for item in evidence
    )


def update_excel(
    excel_path: str,
    summaries: List[Dict[str, Any]],
    selected_question_count: int,
    total_question_count: int,
    seed: int,
) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ImportError("Excel output requires openpyxl. Install it with: pip install openpyxl") from exc

    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "RAG Scores"
        sheet.append(
            [
                "config",
                "questions_used",
                "Page Recall",
                "Answer-Supported Recall",
                "Precision",
                "Groundedness",
                "updated_at",
            ]
        )

    required_columns = [
        "config",
        "questions_used",
        "Page Recall",
        "Answer-Supported Recall",
        "Precision",
        "Groundedness",
        "updated_at",
    ]

    header = [cell.value for cell in sheet[1]]
    for column_name in required_columns:
        if column_name not in header:
            sheet.cell(row=1, column=len(header) + 1).value = column_name
            header.append(column_name)

    row_by_config = {
        sheet.cell(row=row, column=1).value: row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value
    }

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for summary in summaries:
        config = RAG_CONFIGS[summary["config"]]
        row_values = {
            "config": config.name,
            "questions_used": selected_question_count,
            "Page Recall": summary.get(next(key for key in summary if key.startswith("Page Recall@")), 0.0),
            "Answer-Supported Recall": summary.get(
                next(key for key in summary if key.startswith("Answer-Supported Recall@")),
                0.0,
            ),
            "Precision": summary.get(next(key for key in summary if key.startswith("Precision@")), 0.0),
            "Groundedness": summary["Groundedness"],
            "updated_at": updated_at,
        }

        if config.name in row_by_config:
            row = row_by_config[config.name]
            for col, column_name in enumerate(header, start=1):
                sheet.cell(row=row, column=col).value = row_values.get(column_name)
        else:
            sheet.append([row_values.get(column_name) for column_name in header])

    workbook.save(path)
    log_step(f"Excel scores updated: {excel_path}")


def select_sample_results(
    run_outputs: Dict[str, Dict[str, Any]],
    sample_count: int,
    seed: int,
) -> Dict[str, Dict[str, Any]]:
    if not run_outputs:
        return {}

    first_output = next(iter(run_outputs.values()))
    sample_keys = [
        result.get("sample_key") or str(result.get("question_id"))
        for result in first_output["results"]
    ]
    sample_keys = [sample_key for sample_key in sample_keys if sample_key and sample_key != "None"]

    if len(sample_keys) > sample_count:
        rng = random.Random(seed)
        selected_keys = set(rng.sample(sample_keys, sample_count))
    else:
        selected_keys = set(sample_keys)

    sampled_outputs = {}

    for config_name, output in run_outputs.items():
        sampled_output = output.copy()
        sampled_output["results"] = [
            result
            for result in output["results"]
            if (result.get("sample_key") or str(result.get("question_id"))) in selected_keys
        ]
        sampled_outputs[config_name] = sampled_output

    return sampled_outputs


def save_json_results(path: str, run_outputs: Dict[str, Dict[str, Any]], summaries: List[Dict[str, Any]]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    serializable = {
        "summaries": summaries,
        "configs": {
            name: {
                "label": output["config"].label,
                "description": output["config"].description,
            }
            for name, output in run_outputs.items()
        },
        "sample_results": {
            name: output["results"]
            for name, output in run_outputs.items()
        },
    }

    with open(output_path, "w", encoding="utf-8") as file:
        json.dump(serializable, file, ensure_ascii=False, indent=2)

    log_step(f"Detailed JSON results saved: {output_path}")


def markdown_escape(text: Any) -> str:
    return str(text).replace("\r\n", "\n").replace("\r", "\n").strip()


def save_markdown_answers(path: str, run_outputs: Dict[str, Dict[str, Any]], summaries: List[Dict[str, Any]]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "# RAG Sample Output",
        "",
        f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## Summary",
        "",
        "| Config | Questions | Page Recall | Answer-Supported Recall | Precision | Groundedness |",
        "|---|---:|---:|---:|---:|---:|",
    ]

    for summary in summaries:
        page_key = next(key for key in summary if key.startswith("Page Recall@"))
        answer_supported_key = next(key for key in summary if key.startswith("Answer-Supported Recall@"))
        precision_key = next(key for key in summary if key.startswith("Precision@"))
        lines.append(
            "| "
            f"{summary['config']} | "
            f"{summary['questions']} | "
            f"{summary[page_key]:.4f} | "
            f"{summary[answer_supported_key]:.4f} | "
            f"{summary[precision_key]:.4f} | "
            f"{summary['Groundedness']:.4f} |"
        )

    for config_name, output in run_outputs.items():
        config = output["config"]
        lines.extend(
            [
                "",
                f"## {config.name}: {config.label}",
                "",
                config.description,
                "",
            ]
        )

        for result in output["results"]:
            question_label = result.get("sample_key") or result.get("question_id")
            lines.extend(
                [
                    f"### Question {question_label}",
                    "",
                    "**Question**",
                    "",
                    markdown_escape(result["question"]),
                    "",
                    "**RAG Answer**",
                    "",
                    markdown_escape(result["answer"]),
                    "",
                    "**Standard Answer**",
                    "",
                    markdown_escape(result["standard_answer"]),
                    "",
                    "**Standard Citation**",
                    "",
                    markdown_escape(format_evidence(result["evidence"])),
                    "",
                ]
            )

            if result.get("query_type") or result.get("retrieval_mode"):
                lines.extend(
                    [
                        "**Retrieval Mode**",
                        "",
                        f"- query_type: {markdown_escape(result.get('query_type') or 'N/A')}",
                        f"- retrieval_mode: {markdown_escape(result.get('retrieval_mode') or 'N/A')}",
                        "",
                    ]
                )

            if result["rewritten_queries"]:
                lines.extend(["**Rewritten Queries**", ""])
                for query in result["rewritten_queries"]:
                    lines.append(f"- {markdown_escape(query)}")
                lines.append("")

            lines.extend(["**Retrieved/Reranked Chunks**", ""])
            for rank, doc in enumerate(result["reranked_docs"], start=1):
                metadata = doc.get("metadata", {}) or {}
                lines.append(
                    f"{rank}. {markdown_escape(source_label(metadata))} | {markdown_escape(doc_score(doc))}"
                )
                lines.append("")
                lines.append(markdown_escape(doc.get("text", "")))
            lines.append("")

            lines.extend(["**Evaluation**", ""])
            for key, value in result["evaluation"].items():
                if key == "query":
                    continue
                if isinstance(value, float):
                    lines.append(f"- {key}: {value:.4f}")
                else:
                    lines.append(f"- {key}: {value}")
            lines.append("")

    with open(output_path, "w", encoding="utf-8") as file:
        file.write("\n".join(lines).rstrip() + "\n")

    log_step(f"Markdown answers saved: {output_path}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Unified RAG experiment runner")

    parser.add_argument(
        "--configs",
        default="baseline",
        help="Comma-separated config names or 'all'. Valid: baseline, advanced_chunking, query_rewrite, bm25, hybrid, full.",
    )
    parser.add_argument(
        "--question-set",
        default="set1",
        help="Comma-separated named question sets or 'all'. Valid: set1, set2. Default: set1.",
    )
    parser.add_argument("--questions", default=None, help="Question JSON file. Overrides --question-set.")
    parser.add_argument("--pdf-dir", default="pdf", help="PDF directory.")
    parser.add_argument("--cache-dir", default="./rag_cache", help="Directory for preprocessing and chunk caches.")
    parser.add_argument("--preprocess-cache", default=None, help="Preprocessed PDF cache path. Defaults to <cache-dir>/preprocessed_documents.json.")
    parser.add_argument("--reprocess-pdf", action="store_true", help="Re-read PDFs and overwrite the preprocessing cache.")
    parser.add_argument("--sample-size", type=int, default=5, help="Random sample size. Use 0 for all questions; capped at the selected question set size.")
    parser.add_argument("--seed", type=int, default=42, help="Random sampling seed.")
    parser.add_argument("--sample-output-count", type=int, default=5, help="Number of random evaluated questions to save in sample outputs.")

    parser.add_argument("--api-key", default=None, help=f"LLM API key. Defaults to ${DEFAULT_API_KEY_ENV}.")
    parser.add_argument("--generation-model", default=DEFAULT_GENERATION_MODEL, help="Answer generation model.")
    parser.add_argument("--evaluation-model", default=DEFAULT_EVALUATION_MODEL, help="Evaluation model.")
    parser.add_argument("--rewrite-model", default=DEFAULT_REWRITE_MODEL, help="Query rewrite model.")
    parser.add_argument("--embed-model", default=DEFAULT_EMBED_MODEL, help="Embedding model.")

    parser.add_argument("--chroma-dir", default="./chroma_db", help="Chroma persistence directory.")
    parser.add_argument("--rebuild-index", action="store_true", help="Rebuild required Chroma collections.")

    parser.add_argument("--chunk-size", type=int, default=800, help="Fixed chunk size.")
    parser.add_argument("--chunk-overlap", type=int, default=100, help="Fixed chunk overlap.")
    parser.add_argument("--advanced-min-chunk-size", type=int, default=250, help="Advanced chunking minimum body size.")
    parser.add_argument("--advanced-max-chunk-size", type=int, default=800, help="Advanced chunking maximum chunk size.")
    parser.add_argument("--advanced-sentence-overlap", type=int, default=1, help="Advanced chunking sentence overlap.")

    parser.add_argument("--top-k-retrieve", type=int, default=10, help="Initial retrieval size.")
    parser.add_argument("--top-k-rerank", type=int, default=5, help="Reranked context size.")
    parser.add_argument("--evaluation-k", type=int, default=5, help="k used by recall metrics.")
    parser.add_argument("--rewrite-count", type=int, default=3, help="Number of rewritten queries.")
    parser.add_argument("--rewrite-top-k-per-query", type=int, default=5, help="Top-k per query in rewrite RRF.")
    parser.add_argument("--hybrid-rrf-k", type=int, default=60, help="RRF k for hybrid retrieval.")
    parser.add_argument("--semantic-weight", type=float, default=1.0, help="Hybrid semantic weight.")
    parser.add_argument("--bm25-weight", type=float, default=0.5, help="Hybrid BM25 weight.")
    parser.add_argument("--max-context-chars", type=int, default=None, help="Optional context character limit.")

    parser.add_argument("--excel", default="output/rag_scores.xlsx", help="Excel score table path.")
    parser.add_argument("--json-output", default="output/sample_output.json", help="Detailed JSON output path.")
    parser.add_argument("--markdown-output", default="output/sample_output.md", help="Markdown answer output path.")
    parser.add_argument("--checkpoint", default="output/rag_checkpoint.json", help="Checkpoint path for resuming interrupted runs.")
    parser.add_argument("--no-resume", action="store_true", help="Do not resume from an existing checkpoint.")
    parser.add_argument("--reset-checkpoint", action="store_true", help="Delete the checkpoint before running.")
    parser.add_argument("--no-json", action="store_true", help="Do not write detailed JSON results.")
    parser.add_argument("--no-markdown", action="store_true", help="Do not write Markdown answer results.")
    parser.add_argument("--no-excel", action="store_true", help="Do not write Excel score table.")

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    api_key = args.api_key or os.getenv(DEFAULT_API_KEY_ENV)
    config_names = parse_config_names(args.configs)
    selected_configs = [RAG_CONFIGS[name] for name in config_names]

    question_paths = resolve_question_paths(args.question_set, args.questions)
    all_questions = load_questions_from_paths(question_paths)
    selected_questions = sample_questions(
        questions=all_questions,
        sample_size=args.sample_size,
        seed=args.seed,
    )

    question_sources = ", ".join(f"{name}={path}" for name, path in question_paths)
    all_question_counts = question_set_counts(all_questions)
    selected_question_counts = question_set_counts(selected_questions)
    log_step(
        f"Selected {len(selected_questions)} of {len(all_questions)} questions "
        f"from {question_sources} for configs: {', '.join(config_names)}"
    )
    log_step(
        f"Question set distribution: pool={all_question_counts}, "
        f"sample={selected_question_counts}"
    )

    run_signature = build_run_signature(
        args=args,
        config_names=config_names,
        selected_questions=selected_questions,
        question_paths=question_paths,
    )

    if args.reset_checkpoint:
        checkpoint_path = Path(args.checkpoint)
        if checkpoint_path.exists():
            checkpoint_path.unlink()
            log_step(f"Deleted checkpoint: {args.checkpoint}")

    checkpoint_results: Dict[str, List[Dict[str, Any]]] = {}

    if not args.no_resume:
        checkpoint_results = load_checkpoint_results(
            checkpoint_path=args.checkpoint,
            run_signature=run_signature,
        )

    def checkpoint_callback(config_name: str, config_results: List[Dict[str, Any]]) -> None:
        checkpoint_results[config_name] = config_results
        save_checkpoint_results(
            checkpoint_path=args.checkpoint,
            run_signature=run_signature,
            results_by_config=checkpoint_results,
        )

    preprocess_cache = args.preprocess_cache or str(Path(args.cache_dir) / "preprocessed_documents.json")

    prepared_by_strategy = prepare_chunks_and_indexes(
        selected_configs=selected_configs,
        pdf_dir=args.pdf_dir,
        cache_dir=args.cache_dir,
        preprocess_cache=preprocess_cache,
        reprocess_pdf=args.reprocess_pdf,
        chroma_dir=args.chroma_dir,
        embed_model_name=args.embed_model,
        chunk_size=args.chunk_size,
        chunk_overlap=args.chunk_overlap,
        advanced_min_chunk_size=args.advanced_min_chunk_size,
        advanced_max_chunk_size=args.advanced_max_chunk_size,
        advanced_sentence_overlap=args.advanced_sentence_overlap,
        rebuild=args.rebuild_index,
    )

    run_outputs: Dict[str, Dict[str, Any]] = {}
    summaries: List[Dict[str, Any]] = []
    component_cache: Dict[str, Any] = {}
    reranker = Reranker()

    for config in selected_configs:
        output = run_one_config(
            config=config,
            questions=selected_questions,
            prepared_entry=prepared_by_strategy[config.chunking],
            component_cache=component_cache,
            reranker=reranker,
            api_key=api_key,
            generation_model=args.generation_model,
            evaluation_model=args.evaluation_model,
            rewrite_model=args.rewrite_model,
            embed_model_name=args.embed_model,
            chroma_dir=args.chroma_dir,
            top_k_retrieve=args.top_k_retrieve,
            top_k_rerank=args.top_k_rerank,
            evaluation_k=args.evaluation_k,
            rewrite_count=args.rewrite_count,
            rewrite_top_k_per_query=args.rewrite_top_k_per_query,
            hybrid_rrf_k=args.hybrid_rrf_k,
            semantic_weight=args.semantic_weight,
            bm25_weight=args.bm25_weight,
            max_context_chars=args.max_context_chars,
            existing_results=checkpoint_results.get(config.name, []),
            checkpoint_callback=checkpoint_callback,
        )
        run_outputs[config.name] = output
        summaries.append(output["summary"])

    print("\n" + "=" * 88)
    print("SUMMARY")
    print("=" * 88)
    for summary in summaries:
        print(
            f"{summary['config']}: "
            f"Page Recall={summary[f'Page Recall@{args.evaluation_k}']:.4f}, "
            f"Answer-Supported Recall={summary[f'Answer-Supported Recall@{args.evaluation_k}']:.4f}, "
            f"Precision={summary[f'Precision@{args.evaluation_k}']:.4f}, "
            f"Groundedness={summary['Groundedness']:.4f}"
        )

    if not args.no_excel:
        update_excel(
            excel_path=args.excel,
            summaries=summaries,
            selected_question_count=len(selected_questions),
            total_question_count=len(all_questions),
            seed=args.seed,
        )

    sample_outputs = select_sample_results(
        run_outputs=run_outputs,
        sample_count=args.sample_output_count,
        seed=args.seed,
    )

    if not args.no_json:
        save_json_results(
            path=args.json_output,
            run_outputs=sample_outputs,
            summaries=summaries,
        )

    if not args.no_markdown:
        save_markdown_answers(
            path=args.markdown_output,
            run_outputs=sample_outputs,
            summaries=summaries,
        )


if __name__ == "__main__":
    main()
